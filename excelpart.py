from openpyxl import Workbook
import graphingpart
wb = Workbook()
ws = wb.active

class MAIN:
    def __init__(self):
        None
    def main(self):
        self.create_worksheet()
        self.write_worksheet()
        None
    def create_worksheet(self):
        ws.title = "New Title"
        
    def write_worksheet(self):
        print(graphingpart.main.x)
        ws['A1'] = 'Number'
        ws['B1'] = 'Maximum Number'
        ws['C1'] = 'Steps before loop is reached'
        for i in range(len(graphingpart.main.x)):
            print (i)
            ws.append([graphingpart.main.x[i], graphingpart.main.y[i], graphingpart.main.y2[i]])
        wb.save('test.xlsx')

main = MAIN()
